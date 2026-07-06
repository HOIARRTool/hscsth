import re
import html
from io import BytesIO
from pathlib import Path

import numpy as np
import pandas as pd
import plotly.graph_objects as go
import streamlit as st
import openpyxl


st.set_page_config(
    page_title="HSCS Dashboard",
    page_icon="📊",
    layout="wide"
)

# =========================================================
# Shared thresholds / colors
# =========================================================
H_RED_BG = "#FF2B2B"       # แดงสด
H_ORANGE_BG = "#EF6C00"    # ส้มแก่
H_YELLOW_BG = "#F3E58A"    # เหลืองนวลตา
H_GREEN_BG = "#2E7D32"     # เขียวเข้ม
H_MISSING_BG = "#E8EEF6"   # เทาอ่อนสำหรับช่องไม่มีข้อมูล
H_MISSING_FG = "#64748B"

BASE_DIR = Path(__file__).resolve().parent

HSCS_YEAR_CONFIG = {
    "2568": {
        "label": "ปี 2568",
        "default_sheet": "HSCS2568",
        "upload_label": "อัปโหลดไฟล์ Excel HSCS ปี 2568",
    },
    "2569": {
        "label": "ปี 2569",
        "default_sheet": "HSCS2569",
        "upload_label": "อัปโหลดไฟล์ Excel HSCS ปี 2569",
    },
}

REPORT_URL = "https://sites.google.com/view/mch-hscs67-68/%E0%B8%A0%E0%B8%B2%E0%B8%9E%E0%B8%A3%E0%B8%A7%E0%B8%A1?authuser=0"
REPORT_PREVIEW_IMAGE = BASE_DIR / "hscs_report_preview.png"

HAI_LOGO_URL = "https://github.com/HOIARRTool/appqtbi/blob/main/messageImage_1763018963411.jpg?raw=true"


# =========================================================
# Scoring helpers
# =========================================================
def classify_score(score: float) -> tuple[str, str]:
    """Return status label and color group for a % positive response score."""
    if score < 60:
        return "ควรพัฒนาด่วน", "แดง"
    elif 60 <= score <= 70:
        return "เร่งพัฒนา", "ส้ม"
    elif 70 < score <= 80:
        return "ควรพัฒนาต่อเนื่อง", "เหลือง"
    else:
        return "ควรส่งเสริม", "เขียว"


def heatmap_bg_color(score) -> str:
    if pd.isna(score):
        return H_MISSING_BG
    score = float(score)
    if score < 60:
        return H_RED_BG
    elif 60 <= score <= 70:
        return H_ORANGE_BG
    elif 70 < score <= 80:
        return H_YELLOW_BG
    return H_GREEN_BG


def heatmap_font_color(score) -> str:
    if pd.isna(score):
        return H_MISSING_FG
    score = float(score)
    if score < 60:
        return "#FFFFFF"
    elif 60 <= score <= 70:
        return "#FFFFFF"
    elif 70 < score <= 80:
        return "#111111"
    return "#FFFFFF"


def _score_status(score: float) -> tuple[str, str, str]:
    """Return status label, background color, and text color for a score."""
    if pd.isna(score):
        return "ไม่มีข้อมูล", "#F8FAFC", "#0F172A"
    status, _ = classify_score(float(score))
    bg = heatmap_bg_color(score)
    fg = heatmap_font_color(score)
    return status, bg, fg


def _dimension_sort_key(dim_name: str):
    """Sort dimensions by leading number when available, otherwise by text."""
    m = re.match(r"^\s*(\d+)", str(dim_name))
    if m:
        return (0, int(m.group(1)), str(dim_name))
    return (1, 999, str(dim_name))


def _sub_code_sort_key(code: str):
    """Sort sub codes such as A1, A10, B2 naturally."""
    s = str(code or "")
    m = re.match(r"^([A-Za-z]+)(\d+)$", s)
    if m:
        return (m.group(1), int(m.group(2)))
    return (s, 0)


def dedupe_labels(labels):
    seen = {}
    out = []
    for lab in labels:
        if lab not in seen:
            seen[lab] = 1
            out.append(lab)
        else:
            seen[lab] += 1
            out.append(f"{lab} ({seen[lab]})")
    return out


def get_heatmap_display_mode(unit_count: int) -> dict:
    """
    Control matrix width.

    When many units are displayed, forcing the chart to fit the browser width
    makes each cell too narrow. A fixed wide Plotly canvas keeps numbers legible;
    the user can horizontally scroll / zoom as needed.
    """
    if unit_count <= 1:
        return {"compact": True, "width": 760}
    if unit_count == 2:
        return {"compact": True, "width": 920}
    if unit_count <= 18:
        return {"compact": False, "width": None}

    # Around 40 px per unit keeps the cell text readable in the all-groups view.
    return {"compact": True, "width": max(1450, 220 + unit_count * 42)}


# =========================================================
# Uploaded workbook helpers
# =========================================================
@st.cache_data(show_spinner=False)
def get_excel_sheet_names(excel_bytes: bytes) -> list[str]:
    """Return sheet names from an uploaded HSCS Excel workbook."""
    wb = openpyxl.load_workbook(BytesIO(excel_bytes), read_only=True, data_only=True)
    return list(wb.sheetnames)


def _default_sheet_index(sheet_names: list[str], preferred_sheet: str) -> int:
    """Prefer the configured HSCS sheet name, otherwise fall back to the first sheet."""
    if preferred_sheet in sheet_names:
        return sheet_names.index(preferred_sheet)
    return 0


def _uploaded_payloads(uploaded_year_data: dict) -> tuple[tuple[str, str, bytes, str, str], ...]:
    """Create a stable, cacheable payload for trend loading."""
    payloads = []
    for year in HSCS_YEAR_CONFIG.keys():
        if year in uploaded_year_data:
            cfg = uploaded_year_data[year]
            payloads.append((year, cfg["label"], cfg["bytes"], cfg["sheet"], cfg["filename"]))
    return tuple(payloads)


# =========================================================
# Heatmap workbook loader
# =========================================================
def _resolve_header_value(ws, merge_map, row_num, col_num):
    v = ws.cell(row_num, col_num).value
    if v is None and (row_num, col_num) in merge_map:
        v = merge_map[(row_num, col_num)]
    return v


@st.cache_data(show_spinner=False)
def load_interac_heatmap_excel(excel_bytes: bytes, sheet_name: str) -> tuple[pd.DataFrame, list[str]]:
    """
    Read HSCS interac workbook.

    Expected sheet structure:
    - Row 1: top group
    - Row 2: division
    - Row 3: unit
    - Column A: dimension
    - Column B: sub-item
    - Columns C onward: scores
    """
    raw = pd.read_excel(BytesIO(excel_bytes), sheet_name=sheet_name, header=None)

    wb = openpyxl.load_workbook(BytesIO(excel_bytes), data_only=True)
    ws = wb[sheet_name]

    merge_map = {}
    for mr in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = mr.bounds
        if min_row <= 3:
            top_val = ws.cell(min_row, min_col).value
            for r in range(min_row, max_row + 1):
                for c in range(min_col, max_col + 1):
                    merge_map[(r, c)] = top_val

    data_rows = []
    current_dimension = None

    for r in range(3, len(raw)):  # Excel row 4 onward; pandas is 0-based
        dim = raw.iloc[r, 0] if raw.shape[1] > 0 else None
        sub = raw.iloc[r, 1] if raw.shape[1] > 1 else None

        if pd.notna(dim):
            current_dimension = str(dim).strip()

        numeric_found = False
        for c in range(2, raw.shape[1]):
            val = raw.iloc[r, c]
            if pd.notna(val):
                try:
                    float(val)
                    numeric_found = True
                    break
                except Exception:
                    pass

        if pd.notna(sub) and numeric_found:
            sub_text = str(sub).strip()
            code_match = re.match(r"^([A-Z]\d+)\.\s*", sub_text)
            code = code_match.group(1) if code_match else ""
            full_name = re.sub(r"^[A-Z]\d+\.\s*", "", sub_text).strip()
            data_rows.append((r, {"dimension": current_dimension, "sub_code": code, "sub_name": full_name}))

    if not data_rows:
        raise ValueError("ไม่พบข้อมูล heatmap ในชีตที่เลือก")

    row_indices = [r for r, _ in data_rows]

    score_cols = []
    for c in range(2, raw.shape[1]):
        any_numeric = False
        for r in row_indices:
            val = raw.iloc[r, c]
            if pd.notna(val):
                try:
                    float(val)
                    any_numeric = True
                    break
                except Exception:
                    pass
        if any_numeric:
            score_cols.append(c)

    if not score_cols:
        raise ValueError("ไม่พบคอลัมน์คะแนนในชีตที่เลือก")

    records = []
    groups_found = []

    for r, base in data_rows:
        for c in score_cols:
            col_num = c + 1  # pandas 0-based -> openpyxl 1-based

            top_group = _resolve_header_value(ws, merge_map, 1, col_num)
            division = _resolve_header_value(ws, merge_map, 2, col_num)
            unit = _resolve_header_value(ws, merge_map, 3, col_num)

            top_group = str(top_group).replace("\n", " ").strip() if top_group is not None else ""
            division = str(division).replace("\n", " ").strip() if division is not None else ""
            unit = str(unit).replace("\n", " ").strip() if unit is not None else ""

            if not unit:
                unit = division if division else top_group

            groups_found.append(top_group)

            val = raw.iloc[r, c]
            score = np.nan
            if pd.notna(val):
                try:
                    score = float(val)
                except Exception:
                    score = np.nan

            records.append(
                {
                    "group": top_group,
                    "division": division,
                    "unit": unit,
                    "dimension": base["dimension"],
                    "sub_code": base["sub_code"],
                    "sub_name": base["sub_name"],
                    "score": score,
                    "col_index": c,
                }
            )

    long_df = pd.DataFrame(records)

    ordered_groups = []
    for g in groups_found:
        if g and g not in ordered_groups:
            ordered_groups.append(g)

    return long_df, ordered_groups


# =========================================================
# Raw/export workbook loader
# =========================================================
# Codebook derived from the official processed HSCS interac workbook.
# This lets the app read raw respondent-level exports that contain only A1..F6
# columns, without requiring the dimension names to be present in the uploaded file.
ITEM_METADATA = [
    {"code": "A1", "dimension": "1. การทำงานเป็นทีม", "sub_name": "ในหน่วยงานหรือพื้นที่ที่ท่านทำงานอยู่ ทุกคนทำงานร่วมกันเป็นทีมอย่างมีประสิทธิภาพ", "negative": False},
    {"code": "A8", "dimension": "1. การทำงานเป็นทีม", "sub_name": "ในช่วงเวลาที่มีภาระงานยุ่งมาก หรือยาก บุคลากรในหน่วยงานที่ท่านทำงานอยู่จะช่วยเหลือซึ่งกันและกันอย่างดี", "negative": False},
    {"code": "A9", "dimension": "1. การทำงานเป็นทีม", "sub_name": "ในหน่วยงานที่ท่านทำงานอยู่มีปัญหาเรื่องพฤติกรรมในการทำงานของบุคลากรที่ไม่เคารพซึ่งกันและกัน", "negative": True},
    {"code": "A2", "dimension": "2. บุคลากรและพื้นที่การทำงาน", "sub_name": "ในหน่วยงานหรือพื้นที่นี้มีจำนวนบุคลากรในการทำงานที่เพียงพอที่จะรับมือกับปริมาณงานหรือภาระงานของหน่วยงานที่กำหนด", "negative": False},
    {"code": "A3", "dimension": "2. บุคลากรและพื้นที่การทำงาน", "sub_name": "ในหน่วยงานที่ท่านทำงานมีบุคลากรที่ทำงานจำนวนชั่วโมงต่อเนื่องยาวนานมากกว่าที่เหมาะสมต่อการดูแลผู้ป่วยหรือผู้รับบริการได้ดี", "negative": True},
    {"code": "A5", "dimension": "2. บุคลากรและพื้นที่การทำงาน", "sub_name": "หน่วยงานที่ท่านทำงานอยู่ต้องอาศัยหรือพึ่งพาบุคลากรที่มาช่วยทำงานชั่วคราว บุคลากรที่หมุนเวียนมาทำงาน หรือที่ทำงานเฉพาะช่วงที่กำหนดเป็นครั้งๆ มากเกินไป", "negative": True},
    {"code": "A11", "dimension": "2. บุคลากรและพื้นที่การทำงาน", "sub_name": "หน่วยงานที่ท่านทำงานอยู่มักจะเร่งรีบในการทำงานมากเกินไปจนส่งผลกระทบในเชิงลบต่อความปลอดภัยของผู้ป่วยหรือผู้รับบริการ", "negative": True},
    {"code": "A4", "dimension": "3. การเป็นองค์กรแห่งการเรียนรู้ มีการพัฒนาอย่างต่อเนื่อง", "sub_name": "หน่วยงานที่ท่านทำงานอยู่มีการทบทวนกระบวนการทำงานอย่างสม่ำเสมอเพื่อพิจารณาว่าจำเป็นต้องมีการเปลี่ยนแปลงใดบ้างที่ควรปรับปรุงแก้ไขเพื่อความปลอดภัยของผู้ป่วยหรือผู้รับบริการ", "negative": False},
    {"code": "A12", "dimension": "3. การเป็นองค์กรแห่งการเรียนรู้ มีการพัฒนาอย่างต่อเนื่อง", "sub_name": "หน่วยงานที่ท่านทำงานอยู่มีการเปลี่ยนแปลงเพื่อเพิ่มความปลอดภัยให้กับผู้ป่วย/ผู้รับบริการ และมีการประเมินผลความปลอดภัยของผู้ป่วย/ผู้รับบริการว่าดีขึ้นอย่างไร", "negative": False},
    {"code": "A14", "dimension": "3. การเป็นองค์กรแห่งการเรียนรู้ มีการพัฒนาอย่างต่อเนื่อง", "sub_name": "หน่วยงานที่ท่านทำงานอยู่ปล่อยให้เกิดปัญหาเรื่องความปลอดภัยกับผู้ป่วย/ผู้รับบริการในเรื่องเดิมและเกิดขึ้นซ้ำๆ", "negative": True},
    {"code": "A6", "dimension": "4. การตอบสนองต่อความคลาดเคลื่อน", "sub_name": "ในหน่วยงานที่ท่านทำงานอยู่ บุคลากรรู้สึกว่าความผิดพลาดคลาดเคลื่อนที่อาจเกิดขึ้นในการทำงานจะถูกนำมากล่าวโทษกับตัวบุคคล", "negative": True},
    {"code": "A7", "dimension": "4. การตอบสนองต่อความคลาดเคลื่อน", "sub_name": "เมื่อมีการรายงานเหตุการณ์หรืออุบัติการณ์ในหน่วยงานที่ท่านทำงานอยู่ มักมีการระบุว่าเป็นความผิดหรือตำหนิตัวบุคคลเสมอ", "negative": True},
    {"code": "A10", "dimension": "4. การตอบสนองต่อความคลาดเคลื่อน", "sub_name": "เมื่อบุคลากรทำงานผิดพลาดหรือคลาดเคลื่อน หน่วยงานที่ท่านทำงานอยู่จะมุ่งเน้นที่การเรียนรู้มากกว่าการกล่าวโทษตัวบุคคล", "negative": False},
    {"code": "A13", "dimension": "4. การตอบสนองต่อความคลาดเคลื่อน", "sub_name": "ในหน่วยงานที่ท่านทำงานอยู่ไม่มีการสนับสนุนหรือการดูแลบุคลากรที่เกี่ยวข้องกับอุบัติการณ์ความคลาดเคลื่อน หรือเหตุการณ์ไม่พึงประสงค์เกี่ยวกับความปลอดภัยของผู้ป่วย/ผู้รับบริการ", "negative": True},
    {"code": "B1", "dimension": "5. การสนับสนุนของหัวหน้างาน, ผู้จัดการ หรือทีมนำทางคลินิกในเรื่องความปลอดภัย", "sub_name": "หัวหน้างาน หัวหน้ากลุ่มงาน ผู้จัดการ หัวหน้าทีมการดูแลผู้ป่วย หรือประธานทีมนำทางคลินิกของฉันให้ความสำคัญกับข้อเสนอแนะของบุคลากรและผู้ร่วมงานอย่างจริงจังเพื่อปรับปรุงความปลอดภัยของผู้ป่วย", "negative": False},
    {"code": "B2", "dimension": "5. การสนับสนุนของหัวหน้างาน, ผู้จัดการ หรือทีมนำทางคลินิกในเรื่องความปลอดภัย", "sub_name": "หัวหน้างาน หัวหน้ากลุ่มงาน ผู้จัดการ หัวหน้าทีมการดูแลผู้ป่วย หรือประธานทีมนำทางคลินิกของฉันต้องการให้เราทำงานเร็วขึ้นในช่วงเวลาที่ภาระงานยุ่งยาก ชุลมุนหรือเร่งรีบ แม้ว่าต้องลัดขั้นตอนหรือกระบวนการ", "negative": True},
    {"code": "B3", "dimension": "5. การสนับสนุนของหัวหน้างาน, ผู้จัดการ หรือทีมนำทางคลินิกในเรื่องความปลอดภัย", "sub_name": "หัวหน้างาน หัวหน้ากลุ่มงาน ผู้จัดการ หัวหน้าทีมการดูแลผู้ป่วย หรือประธานทีมนำทางคลินิกของฉันมีการดำเนินการเพื่อตอบและแก้ไขความกังวลเรื่องความปลอดภัยของผู้ป่วย/ผู้รับบริการที่ได้รับทราบด้วยความตั้งใจ", "negative": False},
    {"code": "C1", "dimension": "6. การสื่อสารเรื่องความคลาดเคลื่อน", "sub_name": "บุคลากรได้รับการสื่อสารให้ข้อมูลกรณีเกิดความคลาดเคลื่อนขึ้นในหน่วยงานหรือพื้นที่ปฏิบัติงานของตน", "negative": False},
    {"code": "C2", "dimension": "6. การสื่อสารเรื่องความคลาดเคลื่อน", "sub_name": "เมื่อเกิดข้อผิดพลาดหรือความคลาดเคลื่อนในหน่วยงานที่ทำงานอยู่ บุคลากรจะพูดถึงวิธีป้องกันความคลาดเคลื่อนไม่ให้เกิดขึ้นซ้ำอีก", "negative": False},
    {"code": "C3", "dimension": "6. การสื่อสารเรื่องความคลาดเคลื่อน", "sub_name": "ในหน่วยงานที่ฉันปฏิบัติงานอยู่ บุคลากรได้รับการให้ข้อมูลหรือสื่อสารการเปลี่ยนแปลงที่เกิดขึ้นจากรายงานอุบัติการณ์", "negative": False},
    {"code": "C4", "dimension": "7. การสื่อสารที่เปิดกว้าง", "sub_name": "ในหน่วยงานที่ฉันปฏิบัติงานอยู่ บุคลากรสามารถพูดหรือบอกกล่าวได้ทันที เมื่อพบเห็นบางสิ่งบางอย่างที่อาจจะมีผลกระทบหรือผลเสียต่อการดูแลผู้ป่วยหรือผู้รับบริการ", "negative": False},
    {"code": "C5", "dimension": "7. การสื่อสารที่เปิดกว้าง", "sub_name": "ในหน่วยงานที่ฉันปฏิบัติงานอยู่ เมื่อบุคลากรเห็นผู้ที่มีอำนาจ ตำแหน่ง หรือบทบาทหน้าที่สูงกว่า ทำสิ่งที่ไม่ปลอดภัยสำหรับผู้ป่วย/ผู้รับบริการ บุคลากรสามารถพูดหรือบอกกล่าวได้ทันที", "negative": False},
    {"code": "C6", "dimension": "7. การสื่อสารที่เปิดกว้าง", "sub_name": "เมื่อบุคลากรของหน่วยงานที่ฉันปฏิบัติงานอยู่ มีข้อกังวลเรื่องความปลอดภัยของผู้ป่วย/ผู้รับบริการ สามารถพูดหรือบอกกล่าวทันที โดยผู้ที่มีอำนาจ ตำแหน่ง หรือบทบาทหน้าที่สูงกว่าเปิดกว้างในการรับฟัง", "negative": False},
    {"code": "C7", "dimension": "7. การสื่อสารที่เปิดกว้าง", "sub_name": "ในหน่วยงานที่ฉันปฏิบัติงานอยู่ บุคลากรกลัวที่จะตั้งคำถามเมื่อพบสิ่งที่ไม่ถูกต้องหรือผิดปกติ", "negative": True},
    {"code": "D1", "dimension": "8. การรายงานเหตุการณ์ความปลอดภัยของผู้ป่วย/ผู้รับบริการ", "sub_name": "เมื่อเกิดความผิดพลาดหรือคลาดเคลื่อนขึ้นแต่มีการดักจับหรือแก้ไขได้ก่อนถึงตัวผู้ป่วย มีการรายงานอุบัติการณ์ดังกล่าวบ่อยเพียงใด", "negative": False},
    {"code": "D2", "dimension": "8. การรายงานเหตุการณ์ความปลอดภัยของผู้ป่วย/ผู้รับบริการ", "sub_name": "เมื่อเกิดความผิดพลาดหรือคลาดเคลื่อนถึงตัวผู้ป่วย/ผู้รับบริการและอาจทำให้ได้รับอันตรายแต่ไม่เกิดอันตราย มีการรายงานอุบัติการณ์ดังกล่าวบ่อยเพียงใด", "negative": False},
    {"code": "F1", "dimension": "9. การสนับสนุนจากทีมผู้บริหารของสถานพยาบาลในเรื่องความปลอดภัยของผู้ป่วย /ผู้รับบริการ", "sub_name": "การกระทำหรือการดำเนินการของผู้บริหารสถานพยาบาลแสดงให้เห็นถึงการให้ความสำคัญสูงสุดในเรื่องความปลอดภัยของผู้ป่วย/ผู้รับบริการ", "negative": False},
    {"code": "F2", "dimension": "9. การสนับสนุนจากทีมผู้บริหารของสถานพยาบาลในเรื่องความปลอดภัยของผู้ป่วย /ผู้รับบริการ", "sub_name": "ผู้บริหารสถานพยาบาลจัดสรรทรัพยากรอย่างเพียงพอเพื่อปรับปรุงและเพิ่มความปลอดภัยของผู้ป่วยและผู้รับบริการในระบบบริการ", "negative": False},
    {"code": "F3", "dimension": "9. การสนับสนุนจากทีมผู้บริหารของสถานพยาบาลในเรื่องความปลอดภัยของผู้ป่วย /ผู้รับบริการ", "sub_name": "ผู้บริหารโรงพยาบาลเหมือนจะให้ความสนใจเรื่องความปลอดภัยของผู้ป่วย/ผู้รับบริการ เฉพาะกรณีเกิดเหตุการณ์ไม่พึงประสงค์ขึ้นแล้วเท่านั้น", "negative": True},
    {"code": "F4", "dimension": "10. การส่งต่องานและแลกเปลี่ยนข้อมูล ในการเปลี่ยนผ่านระหว่างหน่วยงานหรือเวร", "sub_name": "เมื่อย้ายผู้ป่วย/ผู้รับบริการจากหน่วยงานหรือพื้นที่หนึ่งไปยังอีกหน่วยงานหรืออีกพื้นที่หนึ่ง ข้อมูลสำคัญของผู้ป่วยและการรักษา มักถูกละทิ้งโดยไม่มีการส่งต่อ", "negative": True},
    {"code": "F5", "dimension": "10. การส่งต่องานและแลกเปลี่ยนข้อมูล ในการเปลี่ยนผ่านระหว่างหน่วยงานหรือเวร", "sub_name": "ระหว่างการเปลี่ยนเวรหรือกะการทำงาน ข้อมูลสำคัญในการดูแลผู้ป่วย/ผู้รับบริการ มักถูกละทิ้งหรือละเลยโดยไม่มีการส่งต่อ", "negative": True},
    {"code": "F6", "dimension": "10. การส่งต่องานและแลกเปลี่ยนข้อมูล ในการเปลี่ยนผ่านระหว่างหน่วยงานหรือเวร", "sub_name": "ระหว่างการเปลี่ยนเวรหรือกะการทำงานมีเวลาเพียงพอในการแลกเปลี่ยนและส่งต่อข้อมูลสำคัญการดูแลผู้ป่วย/ผู้รับบริการ ที่สำคัญทั้งหมด", "negative": False},
]
ITEM_CODES = [m["code"] for m in ITEM_METADATA]
ITEM_CODE_SET = set(ITEM_CODES)


def _clean_cell(value) -> str:
    """Clean Excel text values including non-breaking spaces from exports."""
    if pd.isna(value):
        return ""
    return str(value).replace("\xa0", " ").strip()


def _clean_for_number(value):
    if pd.isna(value):
        return np.nan
    return str(value).replace("\xa0", "").strip()


def _norm_col_name(value) -> str:
    return re.sub(r"[\s\u00a0\.\:;_\-\/]+", "", _clean_cell(value).lower())


def _dedupe_columns(cols: list[str]) -> list[str]:
    seen = {}
    out = []
    for c in cols:
        name = c or "Unnamed"
        if name not in seen:
            seen[name] = 1
            out.append(name)
        else:
            seen[name] += 1
            out.append(f"{name}_{seen[name]}")
    return out


def _find_raw_header_row(raw_preview: pd.DataFrame) -> int | None:
    """Find the row whose cells contain HSCS item codes such as A1 and F6."""
    for i in range(len(raw_preview)):
        values = {_clean_cell(v) for v in raw_preview.iloc[i].tolist()}
        hit_count = len(values.intersection(ITEM_CODE_SET))
        if "A1" in values and hit_count >= 5:
            return i
    return None


def _read_raw_survey_sheet(excel_bytes: bytes, sheet_name: str) -> pd.DataFrame:
    raw = pd.read_excel(BytesIO(excel_bytes), sheet_name=sheet_name, header=None, dtype=object)
    header_row = _find_raw_header_row(raw.head(12))
    if header_row is None:
        raise ValueError("ไม่พบหัวตารางแบบ raw/export ที่มีรหัสข้อ A1, A2, ...")

    headers = []
    for c in range(raw.shape[1]):
        header = _clean_cell(raw.iloc[header_row, c])
        if not header and header_row > 0:
            # Some exports keep demographic headers in the row above the item-code row.
            header = _clean_cell(raw.iloc[header_row - 1, c])
        if not header:
            header = f"Unnamed_{c}"
        headers.append(header)

    df = raw.iloc[header_row + 1:].copy()
    df.columns = _dedupe_columns(headers)
    df = df.dropna(how="all")
    for c in df.columns:
        df[c] = df[c].map(_clean_cell)
    return df


def _find_column(columns, candidates: list[str]) -> str | None:
    norm_map = {_norm_col_name(c): c for c in columns}
    for cand in candidates:
        norm = _norm_col_name(cand)
        if norm in norm_map:
            return norm_map[norm]
    # fallback: partial containment for Thai variants
    for c in columns:
        n = _norm_col_name(c)
        for cand in candidates:
            cn = _norm_col_name(cand)
            if cn and (cn in n or n in cn):
                return c
    return None


def _as_numeric_response(series: pd.Series) -> pd.Series:
    vals = pd.to_numeric(series.map(_clean_for_number), errors="coerce")
    # HSCS item responses are 1-5. Other values are excluded from denominators.
    return vals.where(vals.isin([1, 2, 3, 4, 5]))


def _positive_rate(series: pd.Series, negative: bool) -> float:
    vals = _as_numeric_response(series).dropna()
    if vals.empty:
        return np.nan
    if negative:
        return float(vals.isin([1, 2]).mean() * 100)
    return float(vals.isin([4, 5]).mean() * 100)


def _ordered_unique(values: pd.Series | list) -> list[str]:
    out = []
    for v in values:
        s = _clean_cell(v)
        if not s:
            continue
        if s not in out:
            out.append(s)
    return out


def _prepare_raw_analysis_df(df: pd.DataFrame) -> tuple[pd.DataFrame, str | None, str | None, str | None, list[str]]:
    item_cols = [code for code in ITEM_CODES if code in df.columns]
    if len(item_cols) < 20:
        raise ValueError("พบรหัสข้อ HSCS ไม่ครบพอสำหรับประมวลผล raw/export")

    include_col = _find_column(df.columns, ["Include_In_Analysis", "Include in Analysis", "ใช้วิเคราะห์", "นำมาวิเคราะห์"])
    if include_col:
        include_norm = df[include_col].astype(str).str.strip().str.lower()
        df = df[include_norm.isin(["yes", "y", "true", "1", "ใช่", "ใช้", "include"])].copy()

    # Keep rows with at least one valid item response.
    response_matrix = pd.DataFrame({code: _as_numeric_response(df[code]) for code in item_cols})
    df = df[response_matrix.notna().any(axis=1)].copy()

    unit_col = _find_column(df.columns, ["งาน", "พื้นที่ปฏิบัติงาน", "หน่วยงาน", "หน่วย", "work area", "unit"])
    group_col = _find_column(df.columns, ["กลุ่มตามสรพ.", "กลุ่มตาม สรพ.", "กลุ่มตามสรพ", "กลุ่มตาม สรพ", "HAI group", "group"])
    division_col = _find_column(df.columns, ["กลุ่มงาน", "ฝ่าย", "ฝ่าย/งาน", "division", "department"])

    if unit_col is None:
        df["__unit__"] = "ภาพรวม"
        unit_col = "__unit__"
    if group_col is None:
        df["__group__"] = "ไม่ระบุกลุ่ม"
        group_col = "__group__"
    if division_col is None:
        df["__division__"] = df[group_col].replace("", np.nan).fillna("ไม่ระบุกลุ่ม")
        division_col = "__division__"

    for c, fallback in [(unit_col, "ไม่ระบุหน่วยงาน"), (group_col, "ไม่ระบุกลุ่ม"), (division_col, "ไม่ระบุกลุ่มงาน")]:
        df[c] = df[c].map(_clean_cell).replace("", fallback)

    return df, unit_col, group_col, division_col, item_cols


def _scores_for_subset(subset: pd.DataFrame, item_cols: list[str], group: str, division: str, unit: str, col_index: int) -> list[dict]:
    rows = []
    for meta in ITEM_METADATA:
        code = meta["code"]
        if code not in item_cols:
            continue
        score = _positive_rate(subset[code], bool(meta["negative"]))
        rows.append(
            {
                "group": group,
                "division": division,
                "unit": unit,
                "dimension": meta["dimension"],
                "sub_code": code,
                "sub_name": meta["sub_name"],
                "score": score,
                "col_index": col_index,
            }
        )
    return rows


@st.cache_data(show_spinner=False)
def load_raw_survey_excel(excel_bytes: bytes, sheet_name: str) -> tuple[pd.DataFrame, list[str]]:
    """
    Read respondent-level HSCS export files and convert them into the same long
    heatmap structure used by the processed interac workbooks.
    """
    raw_df = _read_raw_survey_sheet(excel_bytes, sheet_name)
    df, unit_col, group_col, division_col, item_cols = _prepare_raw_analysis_df(raw_df)

    records = []
    col_index = 1

    # Grand total column for the executive overview.
    records.extend(_scores_for_subset(df, item_cols, "ภาพรวม", "ภาพรวม", "ภาพรวม", col_index))
    col_index += 1

    # All work areas across all HAI groups. This powers "ภาพรวมทุกกลุ่ม".
    for unit in _ordered_unique(df[unit_col]):
        subset = df[df[unit_col] == unit]
        records.extend(_scores_for_subset(subset, item_cols, "ภาพรวม", "ภาพรวมทุกกลุ่ม", unit, col_index))
        col_index += 1

    # Work areas within each HAI group. These power the group-specific pages.
    ordered_groups = _ordered_unique(df[group_col])
    for group in ordered_groups:
        gdf = df[df[group_col] == group]
        for unit in _ordered_unique(gdf[unit_col]):
            udf = gdf[gdf[unit_col] == unit]
            divisions = _ordered_unique(udf[division_col])
            division = divisions[0] if divisions else group
            records.extend(_scores_for_subset(udf, item_cols, group, division, unit, col_index))
            col_index += 1

    long_df = pd.DataFrame(records)
    return long_df, ordered_groups


@st.cache_data(show_spinner=False)
def load_heatmap_excel(excel_bytes: bytes, sheet_name: str) -> tuple[pd.DataFrame, list[str]]:
    """
    Auto-detect the uploaded workbook format.
    - Raw/export respondent-level sheet: contains item-code headers A1..F6.
    - Processed interac sheet: contains dimension/sub-item rows and score columns.
    """
    preview = pd.read_excel(BytesIO(excel_bytes), sheet_name=sheet_name, header=None, dtype=object, nrows=12)
    if _find_raw_header_row(preview) is not None:
        return load_raw_survey_excel(excel_bytes, sheet_name)
    return load_interac_heatmap_excel(excel_bytes, sheet_name)


def build_overview_df_from_heatmap(long_df: pd.DataFrame) -> pd.DataFrame:
    """
    Build the executive dashboard source from the 'ภาพรวม' column in HSCS*_interac.xlsx.
    If a workbook has no explicit 'ภาพรวม' column, fallback to mean across units.
    """
    df = long_df.copy()

    overall_mask = (
        df["unit"].astype(str).str.strip().eq("ภาพรวม")
        | df["division"].astype(str).str.strip().eq("ภาพรวม")
        | df["group"].astype(str).str.strip().eq("ภาพรวม")
    )

    if overall_mask.any():
        overall_cols = sorted(df.loc[overall_mask, "col_index"].dropna().unique().tolist())
        target_col = overall_cols[0]
        out = df[df["col_index"] == target_col].copy()
        out = out[["dimension", "sub_code", "sub_name", "score"]].rename(columns={"score": "sub_score"})
    else:
        out = (
            df.groupby(["dimension", "sub_code", "sub_name"], dropna=False)["score"]
            .mean()
            .reset_index()
            .rename(columns={"score": "sub_score"})
        )

    out = out.dropna(subset=["sub_score"]).copy()
    out["sub_score"] = pd.to_numeric(out["sub_score"], errors="coerce")
    out = out.dropna(subset=["sub_score"])

    dim_avg = (
        out.groupby("dimension", dropna=False)["sub_score"]
        .mean()
        .rename("dimension_avg")
        .reset_index()
    )
    out = out.merge(dim_avg, on="dimension", how="left")
    out["development_level"] = out["sub_score"].apply(lambda x: classify_score(float(x))[0])
    return out


# =========================================================
# Dashboard overview page
# =========================================================
def _render_dashboard_css():
    st.markdown(
        """
        <style>
        .hscs-hero {
            background: linear-gradient(135deg, #ffffff 0%, #f4f8ff 100%);
            border: 1px solid #dbe5f0;
            border-radius: 22px;
            padding: 18px 22px 18px 24px;
            margin-bottom: 18px;
            box-shadow: 0 8px 24px rgba(15, 23, 42, 0.045);
            display: flex;
            align-items: center;
            justify-content: space-between;
            gap: 18px;
        }
        .hscs-hero-text { min-width: 0; }
        .hscs-hero h1 {
            color: #173B71;
            margin: 0 0 4px 0;
            font-size: 2.0rem;
            line-height: 1.15;
        }
        .hscs-hero p {
            color: #64748B;
            margin: 0;
            font-size: 1.0rem;
        }
        .hscs-hero-logos {
            display: flex;
            align-items: center;
            justify-content: flex-end;
            gap: 12px;
            flex: 0 0 auto;
        }
        .hscs-hero-logo {
            height: 58px;
            max-width: 155px;
            object-fit: contain;
            background: #FFFFFF;
            border: 1px solid #E2E8F0;
            border-radius: 14px;
            padding: 6px 8px;
            box-shadow: 0 4px 12px rgba(15, 23, 42, 0.08);
        }
        @media (max-width: 760px) {
            .hscs-hero { align-items: flex-start; flex-direction: column; }
            .hscs-hero-logos { justify-content: flex-start; }
            .hscs-hero-logo { height: 48px; max-width: 128px; }
        }
        .hscs-section-title {
            color: #173B71;
            font-weight: 800;
            font-size: 1.35rem;
            margin: 18px 0 10px 0;
            border-left: 5px solid #D7A928;
            padding-left: 12px;
        }
        .hscs-dim-grid {
            display: grid;
            grid-template-columns: repeat(5, minmax(0, 1fr));
            gap: 3px;
            background: #CBD5E1;
            border: 1px solid #CBD5E1;
            border-radius: 14px;
            overflow: hidden;
            box-shadow: 0 10px 28px rgba(15, 23, 42, 0.055);
            margin-bottom: 16px;
        }
        .hscs-dim-tile {
            min-height: 176px;
            padding: 13px 14px 12px 14px;
            display: flex;
            flex-direction: column;
            justify-content: space-between;
        }
        .hscs-dim-title {
            font-weight: 800;
            font-size: 0.88rem;
            line-height: 1.28;
            min-height: 43px;
            display: -webkit-box;
            -webkit-line-clamp: 2;
            -webkit-box-orient: vertical;
            overflow: hidden;
        }
        .hscs-dim-score {
            text-align: center;
            font-weight: 900;
            font-size: 1.95rem;
            line-height: 1.05;
            margin: 4px 0 2px 0;
        }
        .hscs-dim-status {
            text-align: center;
            font-weight: 700;
            font-size: 0.76rem;
            opacity: 0.92;
            margin-bottom: 5px;
        }
        .hscs-sub-divider {
            height: 1px;
            background: currentColor;
            opacity: 0.42;
            margin: 4px 0 7px 0;
        }
        .hscs-subgrid {
            display: flex;
            flex-wrap: wrap;
            gap: 5px 4px;
            justify-content: center;
        }
        .hscs-subitem {
            min-width: 31%;
            padding: 3px 4px 4px 4px;
            border-radius: 8px;
            text-align: center;
            line-height: 1.08;
            border: 1px solid rgba(255, 255, 255, 0.78);
            box-shadow: inset 0 0 0 1px rgba(0, 0, 0, 0.08), 0 1px 2px rgba(15, 23, 42, 0.14);
        }
        .hscs-subitem span {
            display: block;
            font-weight: 900;
            font-size: 0.70rem;
            text-transform: uppercase;
        }
        .hscs-subitem strong {
            display: block;
            font-weight: 800;
            font-size: 0.70rem;
        }
        .hscs-legend-inline {
            display: flex;
            flex-wrap: wrap;
            gap: 12px;
            align-items: center;
            margin: 10px 0 18px 0;
            color: #334155;
            font-size: 0.84rem;
            font-weight: 700;
        }
        .hscs-legend-dot {
            display: inline-block;
            width: 14px;
            height: 14px;
            border-radius: 4px;
            margin-right: 5px;
            vertical-align: -2px;
        }
        @media (max-width: 1400px) {
            .hscs-dim-grid { grid-template-columns: repeat(3, minmax(0, 1fr)); }
        }
        @media (max-width: 900px) {
            .hscs-dim-grid { grid-template-columns: repeat(1, minmax(0, 1fr)); }
            .hscs-dim-tile { min-height: 150px; }
        }
        .hscs-trend-note {
            color: #64748B;
            font-size: 0.88rem;
            margin: -4px 0 14px 0;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )


def _dimension_key(dim_name: str) -> str:
    """Stable key for comparing dimensions across years; use the leading number when present."""
    m = re.match(r"^\s*(\d+)", str(dim_name or ""))
    return m.group(1) if m else str(dim_name or "").strip()


@st.cache_data(show_spinner=False)
def load_dimension_trend_data(uploaded_payloads: tuple[tuple[str, str, bytes, str, str], ...]) -> tuple[pd.DataFrame, list[str]]:
    """
    Load dimension-level overview scores for every uploaded year.

    The trend section uses the same workbook loader and the same
    build_overview_df_from_heatmap() logic as the executive dashboard. This keeps
    the trend section aligned with the dashboard's 'ภาพรวม' scores and prevents
    accidental mixing with group-split columns.
    """
    rows = []
    notes = []

    for year, year_label, excel_bytes, sheet_name, filename in uploaded_payloads:
        try:
            long_df, _ = load_heatmap_excel(excel_bytes, sheet_name=sheet_name)
            overview_df = build_overview_df_from_heatmap(long_df)
        except Exception as exc:
            notes.append(f"โหลดข้อมูล {year_label} จากไฟล์ {filename} ไม่สำเร็จ: {exc}")
            continue

        dim_df = (
            overview_df[["dimension", "dimension_avg"]]
            .drop_duplicates()
            .dropna(subset=["dimension_avg"])
            .copy()
        )
        dim_df["year"] = int(year)
        dim_df["year_label"] = year_label
        dim_df["dimension_key"] = dim_df["dimension"].map(_dimension_key)

        rows.extend(dim_df.to_dict("records"))

    trend_df = pd.DataFrame(rows)
    if trend_df.empty:
        return trend_df, notes

    # Prefer the newest available dimension label for display, but keep numeric ordering.
    latest_labels = (
        trend_df.sort_values(["dimension_key", "year"])
        .groupby("dimension_key", as_index=False)
        .tail(1)[["dimension_key", "dimension"]]
        .rename(columns={"dimension": "display_dimension"})
    )
    trend_df = trend_df.merge(latest_labels, on="dimension_key", how="left")
    return trend_df, notes

def build_dimension_trend_figure(dim_trend_df: pd.DataFrame, dim_label: str) -> go.Figure:
    """Small per-dimension year trend chart for the dashboard."""
    d = dim_trend_df.sort_values("year").copy()

    y_values = pd.to_numeric(d["dimension_avg"], errors="coerce").dropna().tolist()
    if y_values:
        y_min = max(0, (int(min(y_values) // 10) * 10) - 10)
        y_max = min(100, (int(max(y_values) // 10) * 10) + 20)
        if y_max - y_min < 30:
            y_min = max(0, y_min - 10)
            y_max = min(100, y_max + 10)
    else:
        y_min, y_max = 0, 100

    years = d["year"].astype(int).tolist()
    scores = pd.to_numeric(d["dimension_avg"], errors="coerce").tolist()

    fig = go.Figure()
    fig.add_trace(
        go.Scatter(
            x=years,
            y=scores,
            mode="lines+markers+text",
            line=dict(width=2.5, color="#173B71"),
            marker=dict(
                size=11,
                color=[heatmap_bg_color(v) for v in scores],
                line=dict(color="#FFFFFF", width=1.5),
            ),
            text=[f"{v:.1f}%" if pd.notna(v) else "" for v in scores],
            textposition="top center",
            textfont=dict(size=11, color="#0F172A"),
            hovertemplate="ปี %{x}<br>คะแนนเฉลี่ยรายมิติ: %{y:.1f}%<extra></extra>",
            showlegend=False,
        )
    )

    # Show the planning horizon even if future-year data are not available yet.
    # The plotted line still uses only available data points, while the x-axis
    # leaves visual space for future HSCS cycles.
    all_years = [2568, 2569, 2570, 2571, 2572]
    x_min = 2567.75
    x_max = 2572.25

    fig.update_layout(
        title=dict(text=dim_label, font=dict(size=15, color="#34138B"), x=0.0, xanchor="left"),
        paper_bgcolor="#FFFFFF",
        plot_bgcolor="#FFFFFF",
        height=255,
        margin=dict(l=34, r=18, t=58, b=34),
    )
    fig.update_xaxes(
        tickmode="array",
        tickvals=all_years,
        range=[x_min, x_max],
        showgrid=False,
        zeroline=False,
        tickfont=dict(size=11),
    )
    fig.update_yaxes(
        range=[y_min, y_max],
        showgrid=True,
        gridcolor="#E5E7EB",
        zeroline=False,
        tickfont=dict(size=11),
    )
    return fig


def render_dimension_trend_section(uploaded_payloads: tuple[tuple[str, str, bytes, str, str], ...]):
    """Render year-to-year dimension trends, independent of the selected dashboard year."""
    st.markdown('<div class="hscs-section-title">แนวโน้มคะแนนเฉลี่ยรายมิติ</div>', unsafe_allow_html=True)
    st.markdown(
        '<div class="hscs-trend-note">เปรียบเทียบคะแนนเฉลี่ยรายมิติจากคอลัมน์ “ภาพรวม” ของแต่ละปี โดยแสดงแกนเวลา 2568–2572 เพื่อใช้ติดตามแนวโน้มระยะต่อไป</div>',
        unsafe_allow_html=True,
    )

    trend_df, notes = load_dimension_trend_data(uploaded_payloads)
    if trend_df.empty:
        st.info("ยังไม่พบข้อมูลเพียงพอสำหรับแสดงแนวโน้มรายมิติ")
        if notes:
            with st.expander("รายละเอียดการโหลดข้อมูลแนวโน้ม", expanded=False):
                for note in notes:
                    st.write(f"- {note}")
        return

    dim_order = (
        trend_df[["dimension_key", "display_dimension"]]
        .drop_duplicates()
        .sort_values("display_dimension", key=lambda s: s.map(_dimension_sort_key))
    )

    cols_per_row = 4
    dims = dim_order.to_dict("records")
    for start in range(0, len(dims), cols_per_row):
        cols = st.columns(cols_per_row)
        for i, dim_info in enumerate(dims[start:start + cols_per_row]):
            dim_key = dim_info["dimension_key"]
            dim_label = dim_info["display_dimension"]
            dim_trend = trend_df[trend_df["dimension_key"] == dim_key].copy()
            fig = build_dimension_trend_figure(dim_trend, dim_label)
            with cols[i]:
                st.plotly_chart(fig, use_container_width=True, key=f"trend_dim_{dim_key}")

    if notes:
        with st.expander("หมายเหตุการโหลดข้อมูลแนวโน้ม", expanded=False):
            for note in notes:
                st.write(f"- {note}")


def render_overview_dashboard_page(excel_bytes: bytes, heatmap_sheet: str, year_label: str, uploaded_payloads: tuple[tuple[str, str, bytes, str, str], ...]):
    """Executive dashboard using the selected HSCS interac workbook."""
    _render_dashboard_css()

    long_df, _ = load_heatmap_excel(excel_bytes, sheet_name=heatmap_sheet)
    df = build_overview_df_from_heatmap(long_df)

    overall_score = float(df["sub_score"].mean()) if not df.empty else np.nan
    overall_status, _, _ = _score_status(overall_score)
    urgent_count = int((df["sub_score"] < 60).sum())
    orange_count = int(((df["sub_score"] >= 60) & (df["sub_score"] <= 70)).sum())
    dim_count = int(df["dimension"].nunique())
    sub_count = int(df[["sub_code", "sub_name"]].drop_duplicates().shape[0])

    st.markdown(
        f'<div class="hscs-hero"><div class="hscs-hero-text"><h1>HSCS Dashboard</h1>'
        f'<p>Hospital Safety Culture Survey: executive overview + drill-down Color-coded Matrix | {html.escape(year_label)}</p></div></div>',
        unsafe_allow_html=True,
    )

    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Overall Positive Score", f"{overall_score:.1f}%", overall_status)
    m2.metric("จำนวนมิติหลัก", f"{dim_count:,}")
    m3.metric("จำนวนมิติย่อย", f"{sub_count:,}")
    m4.metric("ข้อควรพัฒนาด่วน", f"{urgent_count:,}", f"เร่งพัฒนา {orange_count:,} ข้อ")

    st.markdown('<div class="hscs-section-title">ร้อยละคำตอบเชิงบวก (% Positive Response) จำแนกตามมิติ</div>', unsafe_allow_html=True)

    dim_avg_order = (
        df[["dimension", "dimension_avg"]]
        .drop_duplicates()
        .sort_values("dimension", key=lambda s: s.map(_dimension_sort_key))
    )

    tile_html_parts = []
    for _, dim_row in dim_avg_order.iterrows():
        dim = dim_row["dimension"]
        dim_avg = float(dim_row["dimension_avg"])
        status, bg, fg = _score_status(dim_avg)
        dim_safe = html.escape(str(dim))

        sub_df = df[df["dimension"] == dim].copy()
        sub_df = sub_df.sort_values("sub_code", key=lambda s: s.map(_sub_code_sort_key))

        sub_items = []
        for _, r in sub_df.iterrows():
            code = html.escape(str(r["sub_code"] or ""))
            sub_name = html.escape(str(r["sub_name"] or ""))
            score = float(r["sub_score"])
            sub_status, sub_bg, sub_fg = _score_status(score)
            sub_items.append(
                f'<div class="hscs-subitem" style="background:{sub_bg}; color:{sub_fg};" title="{code}: {sub_name} | {html.escape(sub_status)}">'
                f'<span>{code}</span><strong>{score:.1f}%</strong></div>'
            )

        sub_items_html = "".join(sub_items)
        tile_html_parts.append(
            f'<div class="hscs-dim-tile" style="background:{bg}; color:{fg};" title="{dim_safe}">'
            f'<div class="hscs-dim-title">{dim_safe}</div>'
            f'<div><div class="hscs-dim-score">{dim_avg:.1f}%</div>'
            f'<div class="hscs-dim-status">{html.escape(status)}</div></div>'
            f'<div><div class="hscs-sub-divider"></div>'
            f'<div class="hscs-subgrid">{sub_items_html}</div></div>'
            f'</div>'
        )

    st.markdown(
        f'<div class="hscs-dim-grid">{"".join(tile_html_parts)}</div>',
        unsafe_allow_html=True,
    )

    st.markdown(
        f"""
        <div class="hscs-legend-inline">
            <span><i class="hscs-legend-dot" style="background:{H_GREEN_BG};"></i>ควรส่งเสริม &gt; 80</span>
            <span><i class="hscs-legend-dot" style="background:{H_YELLOW_BG};"></i>ควรพัฒนาต่อเนื่อง 70.1–80</span>
            <span><i class="hscs-legend-dot" style="background:{H_ORANGE_BG};"></i>เร่งพัฒนา 60–70</span>
            <span><i class="hscs-legend-dot" style="background:{H_RED_BG};"></i>ควรพัฒนาด่วน &lt; 60</span>
        </div>
        """,
        unsafe_allow_html=True,
    )

    render_dimension_trend_section(uploaded_payloads)

    st.markdown('<div class="hscs-section-title">Priority list: ข้อที่มีคะแนนต่ำสุด</div>', unsafe_allow_html=True)
    priority = (
        df.sort_values(["sub_score", "dimension", "sub_code"], ascending=[True, True, True])
        .head(12)
        .rename(
            columns={
                "dimension": "มิติหลัก",
                "sub_code": "รหัส",
                "sub_name": "ชื่อมิติย่อย",
                "sub_score": "% Positive Score",
                "development_level": "ระดับการพัฒนา",
            }
        )
    )
    priority["% Positive Score"] = priority["% Positive Score"].map(lambda x: f"{float(x):.1f}%")
    st.dataframe(
        priority[["มิติหลัก", "รหัส", "ชื่อมิติย่อย", "% Positive Score", "ระดับการพัฒนา"]],
        use_container_width=True,
        hide_index=True,
    )


# =========================================================
# Color-coded Matrix page
# =========================================================
def build_heatmap_figure(long_df: pd.DataFrame, title_text: str = "") -> go.Figure:
    df = long_df.copy()

    row_order = df[["sub_code", "sub_name", "dimension"]].drop_duplicates()
    row_order["row_label"] = row_order["sub_code"].replace("", np.nan).fillna("NA")

    col_order = (
        df[["col_index", "unit", "division", "group"]]
        .drop_duplicates()
        .sort_values("col_index")
        .reset_index(drop=True)
    )
    col_order["col_label"] = dedupe_labels(col_order["unit"].tolist())

    df = df.merge(col_order[["col_index", "col_label"]], on="col_index", how="left")

    row_labels = row_order["row_label"].tolist()
    col_labels = col_order["col_label"].tolist()

    pivot = (
        df.assign(row_label=df["sub_code"].replace("", np.nan).fillna("NA"))
        .pivot_table(index="row_label", columns="col_label", values="score", aggfunc="mean")
        .reindex(index=row_labels, columns=col_labels)
    )

    row_meta = row_order.set_index("row_label")[["sub_code", "sub_name", "dimension"]]
    col_meta = col_order.set_index("col_label")[["unit", "division", "group"]]

    customdata = []
    text_x = []
    text_y = []
    text_values = []
    text_colors = []

    for rlab in pivot.index:
        row_cd = []
        for clab in pivot.columns:
            score = pivot.loc[rlab, clab]
            row_cd.append([
                row_meta.loc[rlab, "sub_code"],
                row_meta.loc[rlab, "sub_name"],
                row_meta.loc[rlab, "dimension"],
                col_meta.loc[clab, "unit"],
                col_meta.loc[clab, "division"],
                col_meta.loc[clab, "group"],
            ])

            if pd.notna(score):
                text_x.append(clab)
                text_y.append(rlab)
                text_values.append(f"{score:.1f}")
                text_colors.append(heatmap_font_color(score))

        customdata.append(row_cd)

    z = pivot.values.astype(float)

    colorscale = [
        [0.0, H_RED_BG], [0.599999, H_RED_BG],
        [0.6, H_ORANGE_BG], [0.7, H_ORANGE_BG],
        [0.700001, H_YELLOW_BG], [0.8, H_YELLOW_BG],
        [0.800001, H_GREEN_BG], [1.0, H_GREEN_BG],
    ]

    fig = go.Figure()

    # Render missing values as a soft grey layer underneath the main heatmap.
    # This prevents blank cells from looking like a display error while keeping
    # them visually distinct from true 0% scores, which remain red.
    missing_mask = np.isnan(z)
    if missing_mask.any():
        missing_z = np.where(missing_mask, 1, np.nan)
        fig.add_trace(
            go.Heatmap(
                z=missing_z,
                x=col_labels,
                y=row_labels,
                zmin=0,
                zmax=1,
                colorscale=[[0, H_MISSING_BG], [1, H_MISSING_BG]],
                showscale=False,
                hoverinfo="skip",
                xgap=1,
                ygap=1,
            )
        )

    fig.add_trace(
        go.Heatmap(
            z=z,
            x=col_labels,
            y=row_labels,
            zmin=0,
            zmax=100,
            colorscale=colorscale,
            showscale=False,
            customdata=customdata,
            hovertemplate=(
                "<b>%{customdata[0]}</b><br>"
                "มิติย่อย: %{customdata[1]}<br>"
                "มิติหลัก: %{customdata[2]}<br>"
                "หน่วยงาน: %{customdata[3]}<br>"
                "ฝ่าย/งาน: %{customdata[4]}<br>"
                "กลุ่มงาน: %{customdata[5]}<br>"
                "คะแนน: %{z:.1f}%<extra></extra>"
            ),
            xgap=1,
            ygap=1,
        )
    )

    fig.add_trace(
        go.Scatter(
            x=text_x,
            y=text_y,
            mode="text",
            text=text_values,
            textfont=dict(size=11, color=text_colors),
            hoverinfo="skip",
            showlegend=False,
        )
    )

    # Optional dash marks for cells with no valid denominator / no data.
    missing_x = []
    missing_y = []
    for rlab in pivot.index:
        for clab in pivot.columns:
            if pd.isna(pivot.loc[rlab, clab]):
                missing_x.append(clab)
                missing_y.append(rlab)

    if missing_x:
        fig.add_trace(
            go.Scatter(
                x=missing_x,
                y=missing_y,
                mode="text",
                text=["—"] * len(missing_x),
                textfont=dict(size=11, color=H_MISSING_FG),
                hoverinfo="skip",
                showlegend=False,
            )
        )

    unit_count = len(col_labels)
    display_mode = get_heatmap_display_mode(unit_count)

    fig.update_layout(
        title=None,
        paper_bgcolor="#F8FBFF",
        plot_bgcolor="#F8FBFF",
        margin=dict(l=20, r=20, t=40, b=30),
        height=max(760, 31 * len(row_labels) + 210),
        width=display_mode["width"],
    )

    fig.update_xaxes(title_text="", side="top", tickangle=-35, showgrid=False, tickfont=dict(size=10), automargin=True)
    fig.update_yaxes(title_text="", autorange="reversed", showgrid=False, tickfont=dict(size=11), automargin=True)
    return fig



def _normalize_header_text(series: pd.Series) -> pd.Series:
    """Normalize workbook header values for robust filtering."""
    return series.astype(str).str.replace("\n", " ", regex=False).str.strip()


def select_all_groups_matrix(long_df: pd.DataFrame) -> tuple[pd.DataFrame, str]:
    """
    Select columns for 'Color-coded Matrix: ภาพรวมทุกกลุ่ม'.

    New/rebuilt workbooks can contain an explicit group='ภาพรวม' section with
    one column per unit/work area. Use it when it is truly unit-level.

    Legacy workbooks, especially HSCS2568_interac.xlsx, may contain only one
    overall column named 'ภาพรวม'. In that case, showing only that column makes
    the matrix collapse into one unit. For legacy files, fall back to all
    non-overall unit columns.
    """
    df = long_df.copy()
    overall_tokens = {"ภาพรวม", "", "undefined", "none", "nan"}

    col_headers = (
        df[["col_index", "group", "division", "unit"]]
        .drop_duplicates()
        .copy()
    )
    col_headers["group_norm"] = _normalize_header_text(col_headers["group"])
    col_headers["division_norm"] = _normalize_header_text(col_headers["division"])
    col_headers["unit_norm"] = _normalize_header_text(col_headers["unit"])
    col_headers["unit_norm_lower"] = col_headers["unit_norm"].str.lower()
    col_headers["group_norm_lower"] = col_headers["group_norm"].str.lower()
    col_headers["division_norm_lower"] = col_headers["division_norm"].str.lower()

    # Preferred path for rebuilt 2569 workbook:
    # explicit group='ภาพรวม' with multiple real unit/work-area columns.
    explicit_overall_cols = col_headers[col_headers["group_norm"].eq("ภาพรวม")].copy()
    if not explicit_overall_cols.empty:
        real_overall_cols = explicit_overall_cols[
            ~explicit_overall_cols["unit_norm_lower"].isin(overall_tokens)
        ]
        if real_overall_cols["col_index"].nunique() >= 2:
            selected_cols = real_overall_cols["col_index"].tolist()
            return (
                df[df["col_index"].isin(selected_cols)].copy(),
                "Color-coded Matrix ภาพรวมรวมตามงาน ข้ามทุกกลุ่มตาม สรพ.",
            )

        # If the explicit overall group has several columns, keep them except
        # the single grand-total column where possible.
        if explicit_overall_cols["col_index"].nunique() >= 2:
            selected_cols = real_overall_cols["col_index"].tolist()
            if selected_cols:
                return (
                    df[df["col_index"].isin(selected_cols)].copy(),
                    "Color-coded Matrix ภาพรวมรวมตามงาน ข้ามทุกกลุ่มตาม สรพ.",
                )

    # Legacy path for 2568 workbook:
    # exclude grand-total/overall columns and show every real unit column.
    overall_col_mask = (
        col_headers["unit_norm"].eq("ภาพรวม")
        | col_headers["division_norm"].eq("ภาพรวม")
        | col_headers["group_norm"].eq("ภาพรวม")
    )
    real_unit_mask = ~col_headers["unit_norm_lower"].isin(overall_tokens)
    legacy_cols = col_headers.loc[~overall_col_mask & real_unit_mask, "col_index"].tolist()

    if legacy_cols:
        return (
            df[df["col_index"].isin(legacy_cols)].copy(),
            "Color-coded Matrix ภาพรวมทุกหน่วยงาน ข้ามทุกกลุ่มตาม สรพ.",
        )

    # Last-resort fallback: keep old behavior rather than showing nothing.
    return df.copy(), "Color-coded Matrix ภาพรวมทุกกลุ่ม"


def render_heatmap_page(excel_bytes: bytes, heatmap_sheet: str, selected_page: str, selected_year: str):
    long_df, groups = load_heatmap_excel(excel_bytes, sheet_name=heatmap_sheet)

    if selected_page == "Color-coded Matrix: ภาพรวมทุกกลุ่ม":
        # Use rebuilt all-unit overview columns when available.
        # If an older workbook has only one grand-total 'ภาพรวม' column,
        # fall back to the legacy behavior: show all real unit columns.
        filtered, page_desc = select_all_groups_matrix(long_df)
        page_title = "Color-coded Matrix: ภาพรวมทุกกลุ่ม"
    else:
        target_group = selected_page.replace("Color-coded Matrix: ", "", 1)
        filtered = long_df[long_df["group"] == target_group].copy()
        page_title = f"Color-coded Matrix: {target_group}"
        page_desc = "Color-coded Matrix แยกตามกลุ่มงานจากแถวบนสุด"

    st.title(page_title)
    st.markdown(f"{page_desc} | ปี {selected_year}")

    if filtered.empty:
        st.warning("ไม่มีข้อมูลสำหรับหน้านี้")
        return

    all_dims = filtered["dimension"].dropna().unique().tolist()
    all_units = filtered["unit"].dropna().unique().tolist()

    with st.sidebar.expander("ตัวกรอง Color-coded Matrix", expanded=True):
        dim_filter = st.multiselect(
            "เลือกมิติหลัก",
            options=all_dims,
            default=all_dims,
            key=f"hm_dim_{selected_year}_{selected_page}",
        )
        unit_filter = st.multiselect(
            "เลือกหน่วยงาน/คอลัมน์",
            options=all_units,
            default=all_units,
            key=f"hm_unit_{selected_year}_{selected_page}",
        )

    filtered = filtered[
        filtered["dimension"].isin(dim_filter) &
        filtered["unit"].isin(unit_filter)
    ].copy()

    if filtered.empty:
        st.warning("ไม่มีข้อมูลหลังจากกรอง")
        return

    c1, c2 = st.columns(2)
    c1.metric("จำนวนมิติย่อย", f"{filtered[['sub_code','sub_name']].drop_duplicates().shape[0]:,}")
    c2.metric("จำนวนหน่วยงาน", f"{filtered['unit'].nunique():,}")

    fig = build_heatmap_figure(filtered, title_text="")
    display_mode = get_heatmap_display_mode(filtered["unit"].nunique())

    if display_mode["compact"] and filtered["unit"].nunique() > 18:
        st.caption("มุมมองนี้มีหลายหน่วยงาน จึงแสดงเป็นแผนภาพกว้างขึ้นเพื่อให้อ่านตัวเลขได้ชัดขึ้น สามารถเลื่อนแนวนอนหรือซูมด้วยเครื่องมือของกราฟได้")

    st.plotly_chart(fig, use_container_width=not display_mode["compact"])

    with st.expander("ดูคำอธิบายรหัสมิติย่อย", expanded=False):
        show_map = (
            filtered[["sub_code", "dimension", "sub_name"]]
            .drop_duplicates()
            .sort_values(["dimension", "sub_code", "sub_name"])
            .rename(columns={"sub_code": "รหัส", "dimension": "มิติหลัก", "sub_name": "ชื่อข้อย่อย"})
        )
        st.dataframe(show_map, use_container_width=True, hide_index=True)

# =========================================================
# App shell
# =========================================================
st.sidebar.title("TH-HSCS")

st.sidebar.markdown("### อัปโหลดไฟล์ข้อมูล HSCS")
st.sidebar.caption(
    "อัปโหลดไฟล์ Excel 2 ปีสำหรับใช้งานใน session นี้ รองรับไฟล์ excel โดยไม่ต้องเก็บไฟล์ไว้ใน GitHub repo หรือ redeploy Render"
)

uploaded_year_data = {}
for year, cfg in HSCS_YEAR_CONFIG.items():
    uploaded_file = st.sidebar.file_uploader(
        cfg["upload_label"],
        type=["xlsx", "xlsm"],
        key=f"upload_hscs_{year}",
        help=f"รองรับไฟล์ export/raw ที่มีคอลัมน์ A1..F6 หรือไฟล์ interac เดิม โดย prefer ชีตชื่อ {cfg['default_sheet']}",
    )

    if uploaded_file is None:
        continue

    excel_bytes = uploaded_file.getvalue()
    try:
        sheet_names = get_excel_sheet_names(excel_bytes)
    except Exception as exc:
        st.sidebar.error(f"อ่านรายชื่อชีตจากไฟล์ {uploaded_file.name} ไม่สำเร็จ: {exc}")
        continue

    selected_sheet = st.sidebar.selectbox(
        f"เลือกชีตสำหรับ {cfg['label']}",
        options=sheet_names,
        index=_default_sheet_index(sheet_names, cfg["default_sheet"]),
        key=f"sheet_hscs_{year}",
    )

    uploaded_year_data[year] = {
        "label": cfg["label"],
        "bytes": excel_bytes,
        "sheet": selected_sheet,
        "filename": uploaded_file.name,
    }

selected_year = st.sidebar.selectbox(
    "เลือกปีข้อมูล HSCS",
    options=list(HSCS_YEAR_CONFIG.keys()),
    format_func=lambda y: HSCS_YEAR_CONFIG[y]["label"],
    index=min(1, len(HSCS_YEAR_CONFIG) - 1),
)

if st.sidebar.button("Clear cache / reload data"):
    st.cache_data.clear()
    st.rerun()

st.sidebar.markdown("---")

if not uploaded_year_data:
    st.title("HSCS Dashboard")
    st.info(
        "กรุณาอัปโหลดไฟล์ Excel HSCS อย่างน้อย 1 ไฟล์จากแถบด้านซ้ายก่อนใช้งาน Dashboard และ Color-coded Matrix"
    )
    st.markdown(
        """
        **รูปแบบไฟล์ที่รองรับ**
        - ไฟล์ export/raw จากแบบสอบถาม: มีคอลัมน์ข้อมูลผู้ตอบ เช่น `งาน`, `กลุ่มตามสรพ.` และคอลัมน์ข้อคำถาม `A1` ถึง `F6`
        - ไฟล์ processed/interac เดิม: Row 1–3 เป็นหัวคอลัมน์, Column A เป็นมิติหลัก, Column B เป็นข้อย่อย, Column C onward เป็นคะแนน % Positive Response

        หากอัปโหลดครบ 2 ปี ระบบจะแสดงกราฟแนวโน้มคะแนนเฉลี่ยรายมิติแบบเดิมด้วย
        """
    )
    st.stop()

if selected_year not in uploaded_year_data:
    st.title("HSCS Dashboard")
    st.warning(f"กรุณาอัปโหลดไฟล์สำหรับ {HSCS_YEAR_CONFIG[selected_year]['label']} หรือเลือกปีที่อัปโหลดแล้ว")
    uploaded_labels = ", ".join([uploaded_year_data[y]["label"] for y in uploaded_year_data])
    if uploaded_labels:
        st.caption(f"ปีที่อัปโหลดแล้ว: {uploaded_labels}")
    st.stop()

if len(uploaded_year_data) < len(HSCS_YEAR_CONFIG):
    st.sidebar.warning("อัปโหลดไฟล์ยังไม่ครบ 2 ปี: Dashboard ใช้งานได้ แต่กราฟแนวโน้มจะแสดงเฉพาะปีที่มีข้อมูล")

selected_config = uploaded_year_data[selected_year]
heatmap_source = selected_config["bytes"]
heatmap_sheet = selected_config["sheet"]
uploaded_payloads = _uploaded_payloads(uploaded_year_data)

heatmap_pages = ["Color-coded Matrix: ภาพรวมทุกกลุ่ม"]
try:
    _, group_names = load_heatmap_excel(heatmap_source, sheet_name=heatmap_sheet)
    group_names = [
        g for g in group_names
        if str(g).strip() not in ["", "ภาพรวม", "undefined", "None", "nan"]
    ]
    heatmap_pages += [f"Color-coded Matrix: {g}" for g in group_names]
except Exception as exc:
    st.sidebar.warning(f"โหลดรายชื่อกลุ่มงานไม่ได้: {exc}")

page_options = ["Dashboard ภาพรวม"] + heatmap_pages

page = st.sidebar.radio(
    "เลือกหน้าที่ต้องการดู",
    page_options,
    index=0,
    key=f"page_{selected_year}",
)

st.sidebar.markdown("---")
st.sidebar.markdown(
    """
**เกณฑ์สีที่ใช้ร่วมกัน**
- 🔴 แดง: % Positive Score < 60 = ควรพัฒนาด่วน
- 🟠 ส้ม: % Positive Score 60–70 = เร่งพัฒนา
- 🟡 เหลือง: % Positive Score 70.1–80 = ควรพัฒนาต่อเนื่อง
- 🟢 เขียว: % Positive Score > 80 = ควรส่งเสริม
- ⚪ เทา: ไม่มีข้อมูล / ไม่มีตัวหารที่ใช้คำนวณ
"""
)

if page == "Dashboard ภาพรวม":
    render_overview_dashboard_page(heatmap_source, heatmap_sheet, selected_config["label"], uploaded_payloads)
else:
    render_heatmap_page(heatmap_source, heatmap_sheet, page, selected_year)
